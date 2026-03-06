# SPDX-FileCopyrightText: 2023-present Plane Software, Inc.
# SPDX-License-Identifier: LicenseRef-Plane-Commercial
#
# Licensed under the Plane Commercial License (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# https://plane.so/legals/eula
#
# DO NOT remove or modify this notice.
# NOTICE: Proprietary and confidential. Unauthorized use or distribution is prohibited.

import io
from collections import defaultdict
from datetime import datetime
from urllib.parse import urlparse

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Count, Sum, Q
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from plane.db.models import (
    Issue,
    IssueComment,
    IssueRelation,
    IssueAssignee,
    IssueLabel,
    Cycle,
    CycleIssue,
    Module,
    ModuleIssue,
    Label,
    IssueType,
    FileAsset,
    Project,
)
from plane.ee.models import (
    IssueWorkLog,
    IssueProperty,
    IssuePropertyValue,
)
from plane.settings.storage import S3Storage
from plane.utils.host import base_host


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUMMARY_FONT = Font(bold=True, size=11)


def jira_id(external_id):
    """Extract the trailing Jira ID from a composite external_id like '{projectId}_{resourceId}_{jiraId}'."""
    if not external_id:
        return ""
    return external_id.rsplit("_", 1)[-1]


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


def add_summary_row(ws, label, value, row):
    ws.cell(row=row, column=1, value=label).font = SUMMARY_FONT
    ws.cell(row=row, column=1).fill = SUMMARY_FILL
    ws.cell(row=row, column=2, value=value).font = SUMMARY_FONT
    ws.cell(row=row, column=2).fill = SUMMARY_FILL


class _FakeRequest:
    """Minimal request object so S3Storage can build presigned URLs in a non-HTTP context."""

    def __init__(self):
        web_url = base_host(is_app=True, request=None)
        parsed = urlparse(web_url)
        self.scheme = parsed.scheme or "http"
        self._host = parsed.netloc or "localhost"

    def get_host(self):
        return self._host


class Command(BaseCommand):
    help = "Export import verification data as an XLSX workbook for a given project and external source"

    def add_arguments(self, parser):
        parser.add_argument("--project_id", type=str, required=True, help="UUID of the Plane project")
        parser.add_argument(
            "--source",
            type=str,
            default="JIRA_SERVER",
            help="External source identifier (default: JIRA_SERVER)",
        )
        parser.add_argument(
            "--expires",
            type=int,
            default=7 * 24 * 60 * 60,
            help="Presigned URL expiration in seconds (default: 7 days)",
        )

    def handle(self, *args, **options):
        project_id = options["project_id"]
        source = options["source"]
        expires_in = options["expires"]

        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            raise CommandError(f"Project not found: {project_id}")

        project_identifier = project.identifier

        issues = Issue.objects.filter(project_id=project_id, external_source=source)
        total_issues = issues.count()

        if total_issues == 0:
            raise CommandError(f"No imported issues found for project={project_id}, source={source}")

        self.stdout.write(f"Found {total_issues} imported issues. Building workbook...")

        issue_ids = list(issues.values_list("id", flat=True))

        wb = Workbook()
        wb.remove(wb.active)

        self._sheet_overview(wb, project_id, source, total_issues, issue_ids)
        self._sheet_per_issue(wb, project_id, source, issue_ids, project_identifier)
        self._sheet_issue_types(wb, project_id, source, issues)
        self._sheet_states(wb, project_id, source, issues)
        self._sheet_priorities(wb, project_id, source, issues)
        self._sheet_assignees(wb, project_id, source, issue_ids)
        self._sheet_dates(wb, project_id, source, issues)
        self._sheet_cycles(wb, project_id, source, issue_ids)
        self._sheet_modules(wb, project_id, source, issue_ids)
        self._sheet_labels(wb, project_id, source, issue_ids)
        self._sheet_custom_fields(wb, project_id, source, issue_ids)
        self._sheet_relations_summary(wb, project_id, source, issue_ids, project_identifier)
        self._sheet_relations_detail(wb, project_id, source, issue_ids)
        self._sheet_parent_child_detail(wb, project_id, source, issue_ids, project_identifier)
        self._sheet_attachments_detail(wb, project_id, source, issue_ids)
        self._sheet_attachments_summary(wb, issue_ids)
        self._sheet_worklogs_summary(wb, issue_ids)
        self._sheet_worklogs_detail(wb, issue_ids)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        object_name = f"{project_id}/import_verification_{timestamp}.xlsx"

        self.stdout.write("Uploading to storage...")

        storage = S3Storage(request=None)
        is_uploaded = storage.upload_file(
            file_obj=buffer,
            object_name=object_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if not is_uploaded:
            raise CommandError("Failed to upload file to storage.")

        fake_request = _FakeRequest()
        storage = S3Storage(request=fake_request)
        presigned_url = storage.generate_presigned_url(
            object_name,
            expiration=expires_in,
            http_method="GET",
            disposition="attachment",
            filename=f"import_verification_{project_id[:8]}_{timestamp}.xlsx",
        )

        if not presigned_url:
            raise CommandError("File uploaded but failed to generate presigned URL.")

        self.stdout.write(self.style.SUCCESS(f"Upload complete. Download URL (valid for {expires_in // 3600}h):"))
        self.stdout.write(presigned_url)

    # ──────────────────────────────────────────────
    # Sheet 1: Overview / Summary
    # ──────────────────────────────────────────────
    def _sheet_overview(self, wb, project_id, source, total_issues, issue_ids):
        ws = wb.create_sheet("Overview")
        ws.append(["Metric", "Count"])
        style_header_row(ws, 2)

        comments_total = IssueComment.objects.filter(issue_id__in=issue_ids, external_source=source).count()
        worklogs_total = IssueWorkLog.objects.filter(issue_id__in=issue_ids).count()
        total_tracked = IssueWorkLog.objects.filter(issue_id__in=issue_ids).aggregate(
            total=Coalesce(Sum("duration"), 0)
        )["total"]

        # NOTE: external_source is intentionally omitted — the importer writes
        # different source values on Issue ("JIRA_SERVER") vs FileAsset ("JIRA").
        # Scoping to issue_ids already limits to imported issues.
        attachments_total = FileAsset.objects.filter(
            issue_id__in=issue_ids,
            entity_type=FileAsset.EntityTypeContext.ISSUE_ATTACHMENT,
        ).count()

        relations_total = IssueRelation.objects.filter(
            Q(issue_id__in=issue_ids) | Q(related_issue_id__in=issue_ids),
            project_id=project_id,
        ).count()

        issues_with_start = Issue.objects.filter(id__in=issue_ids, start_date__isnull=False).count()
        issues_with_target = Issue.objects.filter(id__in=issue_ids, target_date__isnull=False).count()
        issues_with_both = Issue.objects.filter(
            id__in=issue_ids, start_date__isnull=False, target_date__isnull=False
        ).count()

        cycles_count = Cycle.objects.filter(project_id=project_id, external_source=source).count()
        modules_count = Module.objects.filter(project_id=project_id, external_source=source).count()
        properties_count = IssueProperty.objects.filter(
            issue_type__project_issue_types__project_id=project_id, external_source=source
        ).count()
        types_count = IssueType.objects.filter(
            project_issue_types__project_id=project_id, external_source=source
        ).count()
        labels_count = Label.objects.filter(project_id=project_id, external_source=source).count()

        unique_assignees = IssueAssignee.objects.filter(issue_id__in=issue_ids).values("assignee_id").distinct().count()

        rows = [
            ("Total Imported Issues", total_issues),
            ("Total Comments", comments_total),
            ("Total Worklogs", worklogs_total),
            ("Total Tracked Time (minutes)", total_tracked),
            ("Total Attachments", attachments_total),
            ("Total Relations", relations_total),
            ("Issues with Start Date", issues_with_start),
            ("Issues with Target Date", issues_with_target),
            ("Issues with Both Dates", issues_with_both),
            ("Cycles Created", cycles_count),
            ("Modules Created", modules_count),
            ("Issue Properties (Custom Fields)", properties_count),
            ("Issue Types", types_count),
            ("Labels", labels_count),
            ("Unique Assignees", unique_assignees),
        ]
        for label, value in rows:
            ws.append([label, value])

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 2: Per-Issue Detail
    # ──────────────────────────────────────────────
    def _sheet_per_issue(self, wb, project_id, source, issue_ids, project_identifier=""):
        ws = wb.create_sheet("Per Issue Detail")

        headers = [
            "Plane Key",
            "Issue ID",
            "Sequence ID",
            "Name",
            "Jira ID",
            "External ID",
            "State",
            "Priority",
            "Issue Type",
            "Start Date",
            "Target Date",
            "Assignees",
            "Labels",
            "Comments",
            "Worklogs",
            "Tracked Time (min)",
            "Attachments",
            "Relations",
            "Parent Jira ID",
            "Parent External ID",
            "Cycles",
            "Modules",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        issues = (
            Issue.objects.filter(id__in=issue_ids).select_related("state", "type", "parent").order_by("sequence_id")
        )

        comment_counts = dict(
            IssueComment.objects.filter(issue_id__in=issue_ids, external_source=source)
            .values("issue_id")
            .annotate(cnt=Count("id"))
            .values_list("issue_id", "cnt")
        )
        worklog_counts = dict(
            IssueWorkLog.objects.filter(issue_id__in=issue_ids)
            .values("issue_id")
            .annotate(cnt=Count("id"))
            .values_list("issue_id", "cnt")
        )
        # duration is already in minutes (importer converts Jira's timeSpentSeconds / 60)
        worklog_durations = dict(
            IssueWorkLog.objects.filter(issue_id__in=issue_ids)
            .values("issue_id")
            .annotate(total=Sum("duration"))
            .values_list("issue_id", "total")
        )

        # NOTE: external_source omitted intentionally (see overview sheet comment)
        attachment_counts = dict(
            FileAsset.objects.filter(
                issue_id__in=issue_ids,
                entity_type=FileAsset.EntityTypeContext.ISSUE_ATTACHMENT,
            )
            .values("issue_id")
            .annotate(cnt=Count("id"))
            .values_list("issue_id", "cnt")
        )

        relation_counts = defaultdict(int)
        for iid, cnt in (
            IssueRelation.objects.filter(issue_id__in=issue_ids)
            .values("issue_id")
            .annotate(cnt=Count("id"))
            .values_list("issue_id", "cnt")
        ):
            relation_counts[iid] += cnt
        for iid, cnt in (
            IssueRelation.objects.filter(related_issue_id__in=issue_ids)
            .values("related_issue_id")
            .annotate(cnt=Count("id"))
            .values_list("related_issue_id", "cnt")
        ):
            relation_counts[iid] += cnt

        assignee_map = defaultdict(list)
        for iid, email in (
            IssueAssignee.objects.filter(issue_id__in=issue_ids)
            .select_related("assignee")
            .values_list("issue_id", "assignee__email")
        ):
            assignee_map[iid].append(email)

        label_map = defaultdict(list)
        for iid, name in (
            IssueLabel.objects.filter(issue_id__in=issue_ids)
            .select_related("label")
            .values_list("issue_id", "label__name")
        ):
            label_map[iid].append(name)

        cycle_map = defaultdict(list)
        for iid, cname in (
            CycleIssue.objects.filter(issue_id__in=issue_ids)
            .select_related("cycle")
            .values_list("issue_id", "cycle__name")
        ):
            cycle_map[iid].append(cname)

        module_map = defaultdict(list)
        for iid, mname in (
            ModuleIssue.objects.filter(issue_id__in=issue_ids)
            .select_related("module")
            .values_list("issue_id", "module__name")
        ):
            module_map[iid].append(mname)

        for issue in issues.iterator(chunk_size=500):
            iid = issue.id
            attachment_count = attachment_counts.get(iid, 0)
            plane_key = f"{project_identifier}-{issue.sequence_id}" if project_identifier else str(issue.sequence_id)
            ws.append(
                [
                    plane_key,
                    str(iid),
                    issue.sequence_id,
                    issue.name[:100],
                    jira_id(issue.external_id),
                    issue.external_id or "",
                    issue.state.name if issue.state else "",
                    issue.priority or "",
                    issue.type.name if issue.type else "",
                    str(issue.start_date) if issue.start_date else "",
                    str(issue.target_date) if issue.target_date else "",
                    ", ".join(assignee_map.get(iid, [])),
                    ", ".join(label_map.get(iid, [])),
                    comment_counts.get(iid, 0),
                    worklog_counts.get(iid, 0),
                    worklog_durations.get(iid, 0),
                    attachment_count,
                    relation_counts.get(iid, 0),
                    jira_id(issue.parent.external_id) if issue.parent else "",
                    issue.parent.external_id if issue.parent else "",
                    ", ".join(cycle_map.get(iid, [])),
                    ", ".join(module_map.get(iid, [])),
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 3: Issue Types Breakdown
    # ──────────────────────────────────────────────
    def _sheet_issue_types(self, wb, project_id, source, issues):
        ws = wb.create_sheet("By Issue Type")
        ws.append(["Issue Type", "Jira ID", "External ID", "Issue Count"])
        style_header_row(ws, 4)

        type_data = issues.values("type__name", "type__external_id").annotate(cnt=Count("id")).order_by("-cnt")
        for row in type_data:
            ws.append(
                [
                    row["type__name"] or "(No Type)",
                    jira_id(row["type__external_id"]),
                    row["type__external_id"] or "",
                    row["cnt"],
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 4: States Breakdown
    # ──────────────────────────────────────────────
    def _sheet_states(self, wb, project_id, source, issues):
        ws = wb.create_sheet("By State")
        ws.append(["State", "State Group", "Issue Count"])
        style_header_row(ws, 3)

        state_data = issues.values("state__name", "state__group").annotate(cnt=Count("id")).order_by("-cnt")
        for row in state_data:
            ws.append(
                [
                    row["state__name"] or "(No State)",
                    row["state__group"] or "",
                    row["cnt"],
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 5: Priorities Breakdown
    # ──────────────────────────────────────────────
    def _sheet_priorities(self, wb, project_id, source, issues):
        ws = wb.create_sheet("By Priority")
        ws.append(["Priority", "Issue Count"])
        style_header_row(ws, 2)

        priority_data = issues.values("priority").annotate(cnt=Count("id")).order_by("-cnt")
        for row in priority_data:
            ws.append([row["priority"] or "(None)", row["cnt"]])

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 6: Assignees Breakdown
    # ──────────────────────────────────────────────
    def _sheet_assignees(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("By Assignee")
        ws.append(["Assignee Email", "Display Name", "Issue Count"])
        style_header_row(ws, 3)

        assignee_data = (
            IssueAssignee.objects.filter(issue_id__in=issue_ids)
            .values("assignee__email", "assignee__display_name")
            .annotate(cnt=Count("issue_id", distinct=True))
            .order_by("-cnt")
        )
        for row in assignee_data:
            ws.append(
                [
                    row["assignee__email"],
                    row["assignee__display_name"] or "",
                    row["cnt"],
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 7: Date Population
    # ──────────────────────────────────────────────
    def _sheet_dates(self, wb, project_id, source, issues):
        ws = wb.create_sheet("Date Population")
        ws.append(["Date Category", "Issue Count"])
        style_header_row(ws, 2)

        total = issues.count()
        has_start = issues.filter(start_date__isnull=False).count()
        has_target = issues.filter(target_date__isnull=False).count()
        has_both = issues.filter(start_date__isnull=False, target_date__isnull=False).count()
        has_neither = issues.filter(start_date__isnull=True, target_date__isnull=True).count()
        has_start_only = issues.filter(start_date__isnull=False, target_date__isnull=True).count()
        has_target_only = issues.filter(start_date__isnull=True, target_date__isnull=False).count()

        rows = [
            ("Total Issues", total),
            ("With Start Date", has_start),
            ("With Target Date", has_target),
            ("With Both Dates", has_both),
            ("Start Date Only", has_start_only),
            ("Target Date Only", has_target_only),
            ("No Dates", has_neither),
        ]
        for label, value in rows:
            ws.append([label, value])

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 8: Cycles
    # ──────────────────────────────────────────────
    def _sheet_cycles(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("Cycles")
        ws.append(["Cycle Name", "Jira ID", "External ID", "Start Date", "End Date", "Issue Count"])
        style_header_row(ws, 6)

        cycles = Cycle.objects.filter(project_id=project_id, external_source=source).order_by("name")
        cycle_issue_counts = dict(
            CycleIssue.objects.filter(issue_id__in=issue_ids)
            .values("cycle_id")
            .annotate(cnt=Count("issue_id"))
            .values_list("cycle_id", "cnt")
        )

        for cycle in cycles:
            ws.append(
                [
                    cycle.name,
                    jira_id(cycle.external_id),
                    cycle.external_id or "",
                    str(cycle.start_date) if cycle.start_date else "",
                    str(cycle.end_date) if cycle.end_date else "",
                    cycle_issue_counts.get(cycle.id, 0),
                ]
            )

        ws.append([])
        add_summary_row(ws, "Total Cycles", cycles.count(), ws.max_row + 1)

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 9: Modules
    # ──────────────────────────────────────────────
    def _sheet_modules(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("Modules")
        ws.append(["Module Name", "Jira ID", "External ID", "Start Date", "Target Date", "Issue Count"])
        style_header_row(ws, 6)

        modules = Module.objects.filter(project_id=project_id, external_source=source).order_by("name")
        module_issue_counts = dict(
            ModuleIssue.objects.filter(issue_id__in=issue_ids)
            .values("module_id")
            .annotate(cnt=Count("issue_id"))
            .values_list("module_id", "cnt")
        )

        for module in modules:
            ws.append(
                [
                    module.name,
                    jira_id(module.external_id),
                    module.external_id or "",
                    str(module.start_date) if module.start_date else "",
                    str(module.target_date) if module.target_date else "",
                    module_issue_counts.get(module.id, 0),
                ]
            )

        ws.append([])
        add_summary_row(ws, "Total Modules", modules.count(), ws.max_row + 1)

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 10: Labels
    # ──────────────────────────────────────────────
    def _sheet_labels(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("Labels")
        ws.append(["Label Name", "Jira ID", "External ID", "Color", "Issue Count"])
        style_header_row(ws, 5)

        labels = Label.objects.filter(project_id=project_id, external_source=source).order_by("name")
        label_issue_counts = dict(
            IssueLabel.objects.filter(issue_id__in=issue_ids)
            .values("label_id")
            .annotate(cnt=Count("issue_id"))
            .values_list("label_id", "cnt")
        )

        for label in labels:
            ws.append(
                [
                    label.name,
                    jira_id(label.external_id),
                    label.external_id or "",
                    label.color or "",
                    label_issue_counts.get(label.id, 0),
                ]
            )

        ws.append([])
        add_summary_row(ws, "Total Labels", labels.count(), ws.max_row + 1)

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet 11: Custom Fields (Issue Properties)
    # ──────────────────────────────────────────────
    def _sheet_custom_fields(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("Custom Fields")
        ws.append(
            [
                "Property Name",
                "Property Type",
                "Jira ID",
                "External ID",
                "Origin",
                "Issue Type",
                "Populated Issue Count",
            ]
        )
        style_header_row(ws, 7)

        properties = (
            IssueProperty.objects.filter(
                issue_type__project_issue_types__project_id=project_id,
                external_source=source,
            )
            .select_related("issue_type")
            .order_by("issue_type__name", "display_name")
        )

        prop_value_counts = dict(
            IssuePropertyValue.objects.filter(issue_id__in=issue_ids, property__external_source=source)
            .values("property_id")
            .annotate(cnt=Count("issue_id", distinct=True))
            .values_list("property_id", "cnt")
        )

        system_as_custom = 0
        for prop in properties:
            jid = jira_id(prop.external_id)
            is_system = jid != "" and not jid.isdigit()
            if is_system:
                system_as_custom += 1
            ws.append(
                [
                    prop.display_name,
                    prop.property_type,
                    jid,
                    prop.external_id or "",
                    "System Field" if is_system else "Custom Field",
                    prop.issue_type.name if prop.issue_type else "",
                    prop_value_counts.get(prop.id, 0),
                ]
            )

        ws.append([])
        add_summary_row(ws, "Total Custom Fields", properties.count(), ws.max_row + 1)
        add_summary_row(ws, "System Fields as Custom", system_as_custom, ws.max_row + 1)

        ws_type = wb.create_sheet("Custom Fields by Type")
        ws_type.append(["Property Type", "Field Count", "Total Populated Values"])
        style_header_row(ws_type, 3)

        type_agg = (
            IssueProperty.objects.filter(
                issue_type__project_issue_types__project_id=project_id,
                external_source=source,
            )
            .values("property_type")
            .annotate(field_count=Count("id"))
            .order_by("property_type")
        )

        type_value_agg = dict(
            IssuePropertyValue.objects.filter(issue_id__in=issue_ids, property__external_source=source)
            .values("property__property_type")
            .annotate(cnt=Count("id"))
            .values_list("property__property_type", "cnt")
        )

        for row in type_agg:
            ptype = row["property_type"]
            ws_type.append(
                [
                    ptype,
                    row["field_count"],
                    type_value_agg.get(ptype, 0),
                ]
            )

        auto_width(ws)
        auto_width(ws_type)

    # ──────────────────────────────────────────────
    # Sheet 12: Relations Summary
    # ──────────────────────────────────────────────
    def _sheet_relations_summary(self, wb, project_id, source, issue_ids, project_identifier=""):
        ws = wb.create_sheet("Relations")
        ws.append(["Relation Type", "Count"])
        style_header_row(ws, 2)

        relation_type_counts = (
            IssueRelation.objects.filter(issue_id__in=issue_ids)
            .values("relation_type")
            .annotate(cnt=Count("id"))
            .order_by("relation_type")
        )

        for row in relation_type_counts:
            ws.append([row["relation_type"], row["cnt"]])

        parent_count = Issue.objects.filter(id__in=issue_ids, parent__isnull=False).count()
        sub_issue_count = Issue.objects.filter(parent_id__in=issue_ids).count()

        ws.append([])
        ws.append(["Parent-Child Relations", ""])
        ws.append(["Issues with Parent", parent_count])
        ws.append(["Issues with Sub-Issues", sub_issue_count])

        ws_detail = wb.create_sheet("Relations Per Issue")
        ws_detail.append(
            [
                "Plane Key",
                "Issue ID",
                "Sequence ID",
                "Name",
                "blocked_by",
                "relates_to",
                "duplicate",
                "start_before",
                "finish_before",
                "implemented_by",
                "Has Parent",
                "Sub-Issue Count",
            ]
        )
        style_header_row(ws_detail, 12)

        relation_types = ["blocked_by", "relates_to", "duplicate", "start_before", "finish_before", "implemented_by"]

        per_issue_forward = defaultdict(lambda: defaultdict(int))
        for iid, rtype, cnt in (
            IssueRelation.objects.filter(issue_id__in=issue_ids)
            .values("issue_id", "relation_type")
            .annotate(cnt=Count("id"))
            .values_list("issue_id", "relation_type", "cnt")
        ):
            per_issue_forward[iid][rtype] += cnt

        per_issue_reverse = defaultdict(lambda: defaultdict(int))
        for iid, rtype, cnt in (
            IssueRelation.objects.filter(related_issue_id__in=issue_ids)
            .values("related_issue_id", "relation_type")
            .annotate(cnt=Count("id"))
            .values_list("related_issue_id", "relation_type", "cnt")
        ):
            per_issue_reverse[iid][rtype] += cnt

        sub_issue_counts = dict(
            Issue.objects.filter(parent_id__in=issue_ids)
            .values("parent_id")
            .annotate(cnt=Count("id"))
            .values_list("parent_id", "cnt")
        )

        issues_for_relations = (
            Issue.objects.filter(id__in=issue_ids)
            .select_related("parent")
            .only("id", "sequence_id", "name", "parent_id")
            .order_by("sequence_id")
        )

        for issue in issues_for_relations.iterator(chunk_size=500):
            iid = issue.id
            forward = per_issue_forward.get(iid, {})
            reverse = per_issue_reverse.get(iid, {})

            plane_key = f"{project_identifier}-{issue.sequence_id}" if project_identifier else str(issue.sequence_id)
            row = [
                plane_key,
                str(iid),
                issue.sequence_id,
                issue.name[:80],
            ]
            for rtype in relation_types:
                row.append(forward.get(rtype, 0) + reverse.get(rtype, 0))
            row.append("Yes" if issue.parent_id else "No")
            row.append(sub_issue_counts.get(iid, 0))
            ws_detail.append(row)

        auto_width(ws)
        auto_width(ws_detail)

    # ──────────────────────────────────────────────
    # Sheet: Relations Detail (row-level)
    # ──────────────────────────────────────────────
    def _sheet_relations_detail(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("Relations Detail")
        headers = [
            "Issue Sequence ID",
            "Issue Jira ID",
            "Issue External ID",
            "Relation Type",
            "Related Issue Sequence ID",
            "Related Issue Jira ID",
            "Related Issue External ID",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        issue_lookup = dict(Issue.objects.filter(id__in=issue_ids).values_list("id", "sequence_id"))
        issue_ext_lookup = dict(Issue.objects.filter(id__in=issue_ids).values_list("id", "external_id"))

        relation_qs = IssueRelation.objects.filter(
            Q(issue_id__in=issue_ids) | Q(related_issue_id__in=issue_ids),
            project_id=project_id,
        )

        all_relation_issue_ids = set()
        for iid, rid in relation_qs.values_list("issue_id", "related_issue_id").iterator(chunk_size=1000):
            all_relation_issue_ids.add(iid)
            all_relation_issue_ids.add(rid)

        non_imported_ids = all_relation_issue_ids - set(issue_ids)
        if non_imported_ids:
            extra_seq = dict(Issue.objects.filter(id__in=non_imported_ids).values_list("id", "sequence_id"))
            extra_ext = dict(Issue.objects.filter(id__in=non_imported_ids).values_list("id", "external_id"))
            issue_lookup.update(extra_seq)
            issue_ext_lookup.update(extra_ext)

        for iid, rid, rtype in relation_qs.values_list(
            "issue_id", "related_issue_id", "relation_type"
        ).iterator(chunk_size=1000):
            ws.append(
                [
                    issue_lookup.get(iid, ""),
                    jira_id(issue_ext_lookup.get(iid)),
                    issue_ext_lookup.get(iid, "") or "",
                    rtype,
                    issue_lookup.get(rid, ""),
                    jira_id(issue_ext_lookup.get(rid)),
                    issue_ext_lookup.get(rid, "") or "",
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet: Parent-Child Detail (row-level)
    # ──────────────────────────────────────────────
    def _sheet_parent_child_detail(self, wb, project_id, source, issue_ids, project_identifier=""):
        ws = wb.create_sheet("Parent-Child Detail")
        headers = [
            "Child Plane Key",
            "Child Sequence ID",
            "Child Jira ID",
            "Child External ID",
            "Parent Plane Key",
            "Parent Sequence ID",
            "Parent Jira ID",
            "Parent External ID",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        children = (
            Issue.objects.filter(id__in=issue_ids, parent__isnull=False)
            .select_related("parent")
            .only("id", "sequence_id", "external_id", "parent__sequence_id", "parent__external_id")
            .order_by("sequence_id")
        )

        pk = project_identifier
        for child in children.iterator(chunk_size=500):
            child_key = f"{pk}-{child.sequence_id}" if pk else str(child.sequence_id)
            parent_key = ""
            if child.parent:
                parent_key = f"{pk}-{child.parent.sequence_id}" if pk else str(child.parent.sequence_id)
            ws.append(
                [
                    child_key,
                    child.sequence_id,
                    jira_id(child.external_id),
                    child.external_id or "",
                    parent_key,
                    child.parent.sequence_id if child.parent else "",
                    jira_id(child.parent.external_id) if child.parent else "",
                    child.parent.external_id if child.parent else "",
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet: Attachments Detail (row-level)
    # ──────────────────────────────────────────────
    def _sheet_attachments_detail(self, wb, project_id, source, issue_ids):
        ws = wb.create_sheet("Attachments Detail")
        headers = [
            "Issue Sequence ID",
            "Issue Jira ID",
            "Issue External ID",
            "Attachment Jira ID",
            "Attachment External ID",
            "Filename",
            "Size",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        issue_seq_lookup = dict(Issue.objects.filter(id__in=issue_ids).values_list("id", "sequence_id"))
        issue_ext_lookup = dict(Issue.objects.filter(id__in=issue_ids).values_list("id", "external_id"))

        file_assets = (
            FileAsset.objects.filter(
                issue_id__in=issue_ids,
                entity_type=FileAsset.EntityTypeContext.ISSUE_ATTACHMENT,
            )
            .only("issue_id", "external_id", "attributes", "size")
            .order_by("issue_id")
        )
        for fa in file_assets.iterator(chunk_size=500):
            attrs = fa.attributes or {}
            issue_ext = issue_ext_lookup.get(fa.issue_id)
            ws.append(
                [
                    issue_seq_lookup.get(fa.issue_id, ""),
                    jira_id(issue_ext),
                    issue_ext or "",
                    jira_id(fa.external_id),
                    fa.external_id or "",
                    attrs.get("name", ""),
                    fa.size or attrs.get("size", ""),
                ]
            )

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet: Attachments Summary (by MIME type)
    # ──────────────────────────────────────────────
    def _sheet_attachments_summary(self, wb, issue_ids):
        ws = wb.create_sheet("Attachments Summary")
        ws.append(["MIME Type", "Count", "Total Size (MB)"])
        style_header_row(ws, 3)

        file_assets = FileAsset.objects.filter(
            issue_id__in=issue_ids,
            entity_type=FileAsset.EntityTypeContext.ISSUE_ATTACHMENT,
        ).only("attributes", "size")

        type_stats = defaultdict(lambda: {"count": 0, "size": 0.0})
        for fa in file_assets.iterator(chunk_size=500):
            attrs = fa.attributes or {}
            mime = attrs.get("type", "unknown")
            if not mime:
                mime = "unknown"
            type_stats[mime]["count"] += 1
            type_stats[mime]["size"] += float(fa.size or attrs.get("size", 0) or 0)

        total_count = 0
        total_size = 0.0
        for mime in sorted(type_stats, key=lambda m: type_stats[m]["size"], reverse=True):
            s = type_stats[mime]
            size_mb = round(s["size"] / (1024 * 1024), 2)
            ws.append([mime, s["count"], size_mb])
            total_count += s["count"]
            total_size += s["size"]

        ws.append([])
        total_mb = round(total_size / (1024 * 1024), 2)
        add_summary_row(ws, "Total Attachments", total_count, ws.max_row + 1)
        add_summary_row(ws, "Total Size (MB)", total_mb, ws.max_row + 1)

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet: Worklogs Summary (by user)
    # ──────────────────────────────────────────────
    def _sheet_worklogs_summary(self, wb, issue_ids):
        ws = wb.create_sheet("Worklogs Summary")
        ws.append(["Logged By (Email)", "Display Name", "Worklog Count", "Total Duration (min)"])
        style_header_row(ws, 4)

        user_stats = (
            IssueWorkLog.objects.filter(issue_id__in=issue_ids)
            .values("logged_by__email", "logged_by__display_name")
            .annotate(cnt=Count("id"), total_duration=Coalesce(Sum("duration"), 0))
            .order_by("-total_duration")
        )

        grand_count = 0
        grand_duration = 0
        for row in user_stats:
            ws.append([
                row["logged_by__email"],
                row["logged_by__display_name"] or "",
                row["cnt"],
                row["total_duration"],
            ])
            grand_count += row["cnt"]
            grand_duration += row["total_duration"]

        ws.append([])
        add_summary_row(ws, "Total Worklogs", grand_count, ws.max_row + 1)
        add_summary_row(ws, "Total Duration (min)", grand_duration, ws.max_row + 1)

        auto_width(ws)

    # ──────────────────────────────────────────────
    # Sheet: Worklogs Detail (row-level)
    # ──────────────────────────────────────────────
    def _sheet_worklogs_detail(self, wb, issue_ids):
        ws = wb.create_sheet("Worklogs Detail")
        headers = [
            "Issue Sequence ID",
            "Issue Jira ID",
            "Logged By (Email)",
            "Display Name",
            "Duration (min)",
            "Description",
            "Created At",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        issue_seq_lookup = dict(
            Issue.objects.filter(id__in=issue_ids)
            .values_list("id", "sequence_id")
        )
        issue_ext_lookup = dict(
            Issue.objects.filter(id__in=issue_ids)
            .values_list("id", "external_id")
        )

        worklogs = (
            IssueWorkLog.objects.filter(issue_id__in=issue_ids)
            .select_related("logged_by")
            .only(
                "issue_id", "logged_by__email", "logged_by__display_name",
                "duration", "description", "created_at",
            )
            .order_by("issue_id", "created_at")
        )
        for wl in worklogs.iterator(chunk_size=500):
            issue_ext = issue_ext_lookup.get(wl.issue_id)
            ws.append([
                issue_seq_lookup.get(wl.issue_id, ""),
                jira_id(issue_ext),
                wl.logged_by.email if wl.logged_by else "",
                wl.logged_by.display_name if wl.logged_by else "",
                wl.duration,
                (wl.description or "")[:200],
                str(wl.created_at) if wl.created_at else "",
            ])

        auto_width(ws)
