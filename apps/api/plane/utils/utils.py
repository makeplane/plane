import os
import json
import re
from io import BytesIO

from django.core.files.uploadedfile import InMemoryUploadedFile
from openpyxl import load_workbook


def split_by_numbering(text):
    # 按照 "1. xxx", "2. xxx" 分割文本，但只匹配纯数字+点开头的行首
    pattern = r'(\d+\.)\s*(.*?)(?=\n\d+\.|\Z)'
    matches = re.findall(pattern, text.strip(), re.DOTALL)
    return {int(num[:-1]): content.strip() for num, content in matches}


def build_description_result_list(description: str, result: str):
    desc_dict = split_by_numbering(description)
    res_dict = split_by_numbering(result)

    result_list = []
    # 遍历所有出现在 description 中的编号
    for key in sorted(desc_dict.keys()):
        description_text = desc_dict[key]
        result_text = res_dict.get(key, "")  # 若无对应结果则返回空字符串
        result_list.append({
            "description": description_text,
            "result": result_text
        })

    return result_list


def parser_excel_case(file_path, sheet_name='case'):
    """使用openpyxl将Excel转换为字典"""

    mapping = {"功能": 'label', '测试内容': 'name', '用例等级': 'priority', '测试目的': 'remark',
               '预置条件': 'precondition', '测试步骤': 'description', '预期结果': 'result','模块':'module'}

    # 加载工作簿
    workbook = load_workbook(file_path)

    # 选择工作表
    if sheet_name:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active  # 默认活动工作表

    # 获取第一行作为列标题
    headers = [(mapping.get(cell.value) or cell.value) for cell in worksheet[1]]

    # 读取数据行
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # 创建字典，跳过空行
        if any(cell is not None for cell in row):
            row_dict = dict(zip(headers, row))
            data.append({key: value for key, value in row_dict.items() if key in mapping.values()})

    return data


def get_extension_without_dot(filename):
    """获取不带点号的扩展名"""
    _, extension = os.path.splitext(filename)
    return extension[1:].lower() if extension else ''


def clear_excel_data(excel_data: list[dict]):
    excel = []
    for data in excel_data:
        description = data.pop('description')
        result = data.pop('result')
        steps = build_description_result_list(description, result)
        data['steps'] = steps
        # 标签
        data['label'] = [label.strip() for label in data['label'].split('\n') if label]

        excel.append(data)
    return excel


def parser_case_file(files: list[InMemoryUploadedFile]) -> list:
    result = list()
    for file in files:
        if (suffix := get_extension_without_dot(file.name)) in ['json']:
            result.extend(json.load(file))
        elif suffix in ['xlsx']:
            data = parser_excel_case(BytesIO(file.read()))
            result.extend(clear_excel_data(data))
        else:
            raise Exception('不是支持的文件类型')
    return result
