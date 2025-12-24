"""
Import/Export System with Pluggable Formatters

Exporter: QuerySet → Serializer → Formatter → File/String
Importer: File/String → Formatter → Serializer → Models
"""

import csv
import json
from abc import ABC, abstractmethod
from io import BytesIO, StringIO
from typing import Any, Dict, List, Union

from openpyxl import Workbook, load_workbook


class BaseFormatter(ABC):
    @abstractmethod
    def encode(self, data: List[Dict]) -> Union[str, bytes]:
        """Data → formatted string/bytes"""
        pass

    @abstractmethod
    def decode(self, content: Union[str, bytes]) -> List[Dict]:
        """Formatted string/bytes → data"""
        pass

    @property
    @abstractmethod
    def extension(self) -> str:
        pass


class JSONFormatter(BaseFormatter):
    def __init__(self, indent: int = 2):
        self.indent = indent

    def encode(self, data: List[Dict]) -> str:
        return json.dumps(data, indent=self.indent, default=str)

    def decode(self, content: str) -> List[Dict]:
        return json.loads(content)

    @property
    def extension(self) -> str:
        return "json"


class CSVFormatter(BaseFormatter):
    def __init__(self, flatten: bool = True, delimiter: str = ",", prettify_headers: bool = True):
        """
        Args:
            flatten: Whether to flatten nested dicts.
            delimiter: CSV delimiter character.
            prettify_headers: If True, transforms 'created_by_name' → 'Created By Name'.
        """
        self.flatten = flatten
        self.delimiter = delimiter
        self.prettify_headers = prettify_headers

    def _prettify_header(self, header: str) -> str:
        """Transform 'created_by_name' → 'Created By Name'"""
        return header.replace("_", " ").title()

    def _normalize_header(self, header: str) -> str:
        """Transform 'Display Name' → 'display_name' (reverse of prettify)"""
        return header.strip().lower().replace(" ", "_")

    def _flatten(self, row: Dict, parent_key: str = "") -> Dict:
        items = {}
        for key, value in row.items():
            new_key = f"{parent_key}__{key}" if parent_key else key
            if isinstance(value, dict):
                items.update(self._flatten(value, new_key))
            elif isinstance(value, list):
                items[new_key] = json.dumps(value)
            else:
                items[new_key] = value
        return items

    def _unflatten(self, row: Dict) -> Dict:
        result = {}
        for key, value in row.items():
            parts = key.split("__")
            current = result
            for part in parts[:-1]:
                current = current.setdefault(part, {})

            if isinstance(value, str):
                try:
                    parsed = json.loads(value)
                    if isinstance(parsed, (list, dict)):
                        value = parsed
                except (json.JSONDecodeError, TypeError):
                    pass

            current[parts[-1]] = value
        return result

    def encode(self, data: List[Dict]) -> str:
        if not data:
            return ""

        if self.flatten:
            data = [self._flatten(row) for row in data]

        # Collect all unique field names in order
        fieldnames = []
        for row in data:
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)

        output = StringIO()

        if self.prettify_headers:
            # Create header mapping: original_key → Pretty Header
            header_map = {key: self._prettify_header(key) for key in fieldnames}
            pretty_headers = [header_map[key] for key in fieldnames]

            # Write pretty headers manually, then write data rows
            writer = csv.writer(output, delimiter=self.delimiter)
            writer.writerow(pretty_headers)

            # Write data rows in the same field order
            for row in data:
                writer.writerow([row.get(key, "") for key in fieldnames])
        else:
            writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=self.delimiter)
            writer.writeheader()
            writer.writerows(data)

        return output.getvalue()

    def decode(self, content: str, normalize_headers: bool = True) -> List[Dict]:
        """
        Decode CSV content to list of dicts.

        Args:
            content: CSV string
            normalize_headers: If True, converts 'Display Name' → 'display_name'
        """
        rows = list(csv.DictReader(StringIO(content), delimiter=self.delimiter))

        # Normalize headers: 'Email' → 'email', 'Display Name' → 'display_name'
        if normalize_headers:
            rows = [{self._normalize_header(k): v for k, v in row.items()} for row in rows]

        if self.flatten:
            rows = [self._unflatten(row) for row in rows]

        return rows

    @property
    def extension(self) -> str:
        return "csv"


class XLSXFormatter(BaseFormatter):
    """Formatter for XLSX (Excel) files using openpyxl."""

    def __init__(self, prettify_headers: bool = True, list_joiner: str = ", "):
        """
        Args:
            prettify_headers: If True, transforms 'created_by_name' → 'Created By Name'.
            list_joiner: String to join list values (default: ", ").
        """
        self.prettify_headers = prettify_headers
        self.list_joiner = list_joiner

    def _prettify_header(self, header: str) -> str:
        """Transform 'created_by_name' → 'Created By Name'"""
        return header.replace("_", " ").title()

    def _normalize_header(self, header: str) -> str:
        """Transform 'Display Name' → 'display_name' (reverse of prettify)"""
        return header.strip().lower().replace(" ", "_")

    def _format_value(self, value: Any) -> Any:
        """Format a value for XLSX cell."""
        if value is None:
            return ""
        if isinstance(value, list):
            return self.list_joiner.join(str(v) for v in value)
        if isinstance(value, dict):
            return json.dumps(value)
        return value

    def encode(self, data: List[Dict]) -> bytes:
        """Encode data to XLSX bytes."""
        wb = Workbook()
        ws = wb.active

        if not data:
            # Return empty workbook
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # Collect all unique field names in order
        fieldnames = []
        for row in data:
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)

        # Write header row
        if self.prettify_headers:
            headers = [self._prettify_header(key) for key in fieldnames]
        else:
            headers = fieldnames
        ws.append(headers)

        # Write data rows
        for row in data:
            ws.append([self._format_value(row.get(key, "")) for key in fieldnames])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def decode(self, content: bytes, normalize_headers: bool = True) -> List[Dict]:
        """
        Decode XLSX bytes to list of dicts.

        Args:
            content: XLSX file bytes
            normalize_headers: If True, converts 'Display Name' → 'display_name'
        """
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # First row is headers
        headers = list(rows[0])
        if normalize_headers:
            headers = [self._normalize_header(str(h)) if h else "" for h in headers]

        # Convert remaining rows to dicts
        result = []
        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    # Try to parse JSON strings back to lists/dicts
                    if isinstance(value, str):
                        try:
                            parsed = json.loads(value)
                            if isinstance(parsed, (list, dict)):
                                value = parsed
                        except (json.JSONDecodeError, TypeError):
                            pass
                    row_dict[headers[i]] = value
            result.append(row_dict)

        return result

    @property
    def extension(self) -> str:
        return "xlsx"
