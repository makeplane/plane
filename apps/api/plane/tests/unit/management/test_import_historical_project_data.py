# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

"""
Integration tests for the "import_historical_project_data" management command
(Phase 3 of the custom-fields roadmap: batch-importing historical rows from the
source spreadsheet into Projects carrying the 23 seeded custom fields).

Coverage complements plane/utils/historical_project_import.py's own pure-function
tests: this file exercises the parts that touch the database -- Project/field
creation, the unique-key duplicate check, --dry-run's rollback, and CommandError
paths -- which the pure module deliberately has no way to cover on its own.
"""

from decimal import Decimal
from io import StringIO

import openpyxl
import pytest
from django.core.management import call_command
from django.core.management.base import CommandError

from plane.db.default_data.project_custom_fields import DEFAULT_PROJECT_CUSTOM_FIELDS
from plane.db.models import Contract, ContractProject, Project, ProjectCustomFieldValue
from plane.tests.factories import ProjectFactory, UserFactory, WorkspaceFactory, WorkspaceMemberFactory

FIELD_NAMES = [spec["name"] for spec in DEFAULT_PROJECT_CUSTOM_FIELDS]

# The full A-W column layout the real source spreadsheet (项目汇总表.xlsx) uses.
# Phase A retired column A ("合同号&项目号" composite) as data, but the layout
# still has 23 columns: 1 retired + 17 project-side fields (DEFAULT_PROJECT_CUSTOM_FIELDS,
# none of which is at column A anymore) + 5 contract-side columns routed into
# Contract / ContractProject. Defining them as an explicit A-W layout (rather
# than zipping DEFAULT_PROJECT_CUSTOM_FIELDS with positions) is the only way to
# keep the worksheet's header row matching the real source file after Phase A's
# schema shuffle.
XLSX_HEADERS = [
    "合同号&项目号",          # col 1 - retired composite, ignored at parse time
    "区域",                    # col 2
    "省份",                    # col 3
    "行业",                    # col 4
    "分支",                    # col 5
    "合同号",                  # col 6  -> Contract.contract_no
    "签约登记日期",             # col 7  -> Contract.sign_date
    "合同净额/不含第三方（人民币万元）",  # col 8  -> Contract.total_amount
    "税率",                   # col 9  -> Contract.tax_rate (source_header differs from display name)
    "合同占比",                # col 10 -> ContractProject.allocation_ratio (source_header differs)
    "客户项目名称",             # col 11
    "项目序号",                # col 12 (is_unique_key=True post-Phase A)
    "公司项目名称",             # col 13
    "项目类别",                # col 14
    "客户域",                  # col 15
    "业务域",                  # col 16
    "生产方式类别",             # col 17
    "公司产品名称",             # col 18
    "核心产品线",              # col 19
    "生产状态",                # col 20
    "验收阶段",                # col 21
    "成本投入状态",             # col 22
    "能否验收状态",             # col 23
]

# Column positions (1-based) of the two percent-formatted cells in the workbook --
# Excel stores a percent-formatted cell as the underlying fraction (13% -> 0.13),
# and coerce_number() only multiplies back up to 13 when the cell's format says
# so. _row() below supplies that fraction.
_PERCENT_COLUMNS = {9, 10}  # 税率 (Contract.tax_rate) and 合同占比 (ContractProject.allocation_ratio)


def _build_workbook(path, rows):
    """rows: list of lists, each the 23 A-W cell values for one data row, written
    starting at row 7 (rows 1-6 hold the group header / field header / notes, same
    layout as the real source spreadsheet)."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    for col_index, header_text in enumerate(XLSX_HEADERS, start=1):
        worksheet.cell(row=4, column=col_index, value=header_text)
    for row_offset, row_values in enumerate(rows):
        for col_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=7 + row_offset, column=col_index, value=value)
            if col_index in _PERCENT_COLUMNS:
                cell.number_format = "0%"
    workbook.save(path)
    return path


def _row(unique_key, customer_name="客户项目名称示例", **overrides):
    """Build one source-spreadsheet row in A-W column order. ``unique_key`` goes
    in column L (项目序号), the new is_unique_key column post-Phase A -- column
    A (the retired "合同号&项目号") is filled with the same value just so legacy
    spreadsheet readers don't show a blank, but the import command now ignores it.
    """
    values = {
        "合同号&项目号": unique_key,            # col 1, ignored by import
        "区域": "华东",                         # col 2
        "省份": "江苏",                         # col 3
        "行业": "金融",                         # col 4
        "分支": "分支A",                       # col 5
        "合同号": "HD0001",                    # col 6 -> Contract.contract_no
        "签约登记日期": "2024-03-15",           # col 7 -> Contract.sign_date
        "合同净额/不含第三方（人民币万元）": 128.5,  # col 8 -> Contract.total_amount
        "税率": 0.13,                          # col 9 -> Contract.tax_rate (is_percent)
        "合同占比": 0.5,                       # col 10 -> ContractProject.allocation_ratio (is_percent)
        "客户项目名称": customer_name,          # col 11
        "项目序号": unique_key,                # col 12 (is_unique_key=True) -- use same key for easy cross-ref
        "公司项目名称": "内部项目名A",          # col 13
        "项目类别": "A",                       # col 14 (dropdown)
        "客户域": "支撑",                      # col 15 (dropdown)
        "业务域": "政企支撑域",                 # col 16 (dropdown)
        "生产方式类别": "Z",                   # col 17 (dropdown)
        "公司产品名称": "账务处理中心",         # col 18 (dropdown)
        "核心产品线": "核心产品线X",           # col 19
        "生产状态": "P3",                      # col 20 (dropdown)
        "验收阶段": "签约",                    # col 21 (dropdown)
        "成本投入状态": "执行（正在投入成本）",  # col 22 (dropdown)
        "能否验收状态": "已知项目（可验收，已签约已交接）",  # col 23 (dropdown)
    }
    values.update(overrides)
    return [values[header] for header in XLSX_HEADERS]


@pytest.fixture
def workspace_and_member(db):
    user = UserFactory()
    workspace = WorkspaceFactory(owner=user)
    WorkspaceMemberFactory(workspace=workspace, member=user, role=20, is_active=True)
    return workspace, user


@pytest.mark.unit
@pytest.mark.django_db
class TestImportHistoricalProjectData:
    def test_imports_row_into_project_and_seeded_fields(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001")])

        out = StringIO()
        call_command(
            "import_historical_project_data",
            str(xlsx_path),
            "--workspace",
            workspace.slug,
            "--created-by",
            user.email,
            stdout=out,
        )

        assert "Created 1 project(s)" in out.getvalue()
        project = Project.objects.get(workspace=workspace)
        assert project.created_by_id == user.id
        assert project.name  # sanitized from 客户项目名称
        assert project.identifier

        unique_value = ProjectCustomFieldValue.objects.get(
            project=project, custom_field__is_unique_key=True
        )
        # Phase A: the is_unique_key flag now points at "项目序号" (column L),
        # not the retired "合同号&项目号" composite.
        assert unique_value.value_text == "HD2024-001&PRJ2024-001"
        assert unique_value.custom_field.name == "项目序号"

        # The percent "税率" cell used to land on ProjectCustomFieldValue; Phase
        # A moved it onto Contract.tax_rate, so assert that path instead.
        contract = Contract.objects.get(workspace=workspace, contract_no="HD0001")
        assert contract.tax_rate == Decimal("13.0000")
        assert contract.total_amount == Decimal("128.5")
        # The 合同占比 percent cell is per-relationship, not per-contract.
        link = ContractProject.objects.get(contract=contract, project=project)
        assert link.allocation_ratio == Decimal("50.0000")

        dropdown_value = ProjectCustomFieldValue.objects.get(project=project, custom_field__name="项目类别")
        assert dropdown_value.value_option.name == "A"

    def test_created_by_email_lookup_is_case_insensitive(self, workspace_and_member, tmp_path):
        # User.save() lowercases email before storing; an operator who types the
        # email with any capitals must still resolve to the same account.
        workspace, user = workspace_and_member
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001")])

        out = StringIO()
        call_command(
            "import_historical_project_data",
            str(xlsx_path),
            "--workspace",
            workspace.slug,
            "--created-by",
            user.email.upper(),
            stdout=out,
        )

        assert "Created 1 project(s)" in out.getvalue()
        assert Project.objects.get(workspace=workspace).created_by_id == user.id

    def test_dry_run_creates_nothing(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001")])

        out = StringIO()
        call_command(
            "import_historical_project_data",
            str(xlsx_path),
            "--workspace",
            workspace.slug,
            "--created-by",
            user.email,
            "--dry-run",
            stdout=out,
        )

        assert "[DRY RUN]" in out.getvalue()
        assert "Created 1 project(s)" in out.getvalue()
        assert Project.objects.filter(workspace=workspace).count() == 0

    def test_rerun_skips_already_imported_unique_key(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001")])

        call_command(
            "import_historical_project_data", str(xlsx_path), "--workspace", workspace.slug, "--created-by", user.email
        )
        out = StringIO()
        call_command(
            "import_historical_project_data",
            str(xlsx_path),
            "--workspace",
            workspace.slug,
            "--created-by",
            user.email,
            stdout=out,
        )

        assert "Created 0 project(s)" in out.getvalue()
        assert "1 duplicate" in out.getvalue()
        assert Project.objects.filter(workspace=workspace).count() == 1

    def test_duplicate_customer_name_gets_disambiguated_project_name(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        xlsx_path = _build_workbook(
            tmp_path / "source.xlsx",
            [
                _row("HD2024-001&PRJ2024-001", customer_name="同名客户"),
                _row("HD2024-002&PRJ2024-002", customer_name="同名客户"),
            ],
        )

        call_command(
            "import_historical_project_data", str(xlsx_path), "--workspace", workspace.slug, "--created-by", user.email
        )

        names = set(Project.objects.filter(workspace=workspace).values_list("name", flat=True))
        assert len(names) == 2  # no UNIQUE constraint violation from the collision

    def test_duplicate_name_at_max_length_does_not_hang(self, workspace_and_member, tmp_path):
        # Regression: f"{base_name}_{suffix}"[:255] alone always collapses back to
        # base_name once base_name is already 255 chars (the suffix falls past the
        # truncation point), so the disambiguation loop never changed the candidate
        # and never terminated. A row whose sanitized customer name is exactly 255
        # chars, colliding with an existing project, must still resolve to a
        # different, valid name -- this test finishing at all is the regression
        # proof; the unfixed code hangs forever on it.
        workspace, user = workspace_and_member
        long_name = "客" * 255
        ProjectFactory(workspace=workspace, name=long_name, identifier="EXISTING1")

        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001", customer_name=long_name)])
        call_command(
            "import_historical_project_data", str(xlsx_path), "--workspace", workspace.slug, "--created-by", user.email
        )

        new_project = Project.objects.exclude(name=long_name).get(workspace=workspace)
        assert new_project.name != long_name
        assert len(new_project.name) <= 255
        assert new_project.name.startswith("客")

    def test_identifiers_do_not_collide_across_many_cjk_only_rows(self, workspace_and_member, tmp_path):
        # Regression: identifier generation used to scan from suffix=1 every row, so
        # many rows that all fall back to the same short ASCII base (every CJK-only
        # name does) turned into an O(rows^2) collision scan. Seeding from row_idx
        # instead should still produce distinct, valid identifiers with no collision,
        # just without the quadratic blowup. This test only asserts correctness
        # (uniqueness); it doesn't assert on query count.
        workspace, user = workspace_and_member
        rows = [_row(f"HD2024-{i:03d}&PRJ2024-{i:03d}", customer_name=f"客户{i}") for i in range(1, 11)]
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", rows)

        out = StringIO()
        call_command(
            "import_historical_project_data",
            str(xlsx_path),
            "--workspace",
            workspace.slug,
            "--created-by",
            user.email,
            stdout=out,
        )

        assert "Created 10 project(s)" in out.getvalue()
        identifiers = list(Project.objects.filter(workspace=workspace).values_list("identifier", flat=True))
        assert len(identifiers) == len(set(identifiers)) == 10

    def test_forbidden_chars_in_customer_name_are_sanitized(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        xlsx_path = _build_workbook(
            tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001", customer_name="某零售(集团)公司&子公司")]
        )

        call_command(
            "import_historical_project_data", str(xlsx_path), "--workspace", workspace.slug, "--created-by", user.email
        )

        project = Project.objects.get(workspace=workspace)
        for forbidden_char in "&()":
            assert forbidden_char not in project.name

    def test_missing_unique_key_row_is_skipped_not_fatal(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        bad_row = _row("HD2024-001&PRJ2024-001")
        # Phase A: blank out column L (项目序号, the new is_unique_key column)
        # rather than column A (the retired 合同号&项目号) which the import
        # command now ignores.
        bad_row[11] = None
        good_row = _row("HD2024-002&PRJ2024-002")
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [bad_row, good_row])

        out = StringIO()
        call_command(
            "import_historical_project_data",
            str(xlsx_path),
            "--workspace",
            workspace.slug,
            "--created-by",
            user.email,
            stdout=out,
        )

        assert "Created 1 project(s)" in out.getvalue()
        assert "1 missing unique key" in out.getvalue()
        assert Project.objects.filter(workspace=workspace).count() == 1

    def test_header_mismatch_raises_command_error(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member
        xlsx_path = tmp_path / "bad_headers.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(row=4, column=1, value="WRONG COLUMN NAME")
        workbook.save(xlsx_path)

        with pytest.raises(CommandError, match="Header row does not match"):
            call_command(
                "import_historical_project_data",
                str(xlsx_path),
                "--workspace",
                workspace.slug,
                "--created-by",
                user.email,
            )

    def test_non_member_created_by_raises_command_error(self, workspace_and_member, tmp_path):
        workspace, _user = workspace_and_member
        # UserFactory (plane/tests/factories.py) does not set `username`, and the
        # model default is "" for every instance, colliding with the fixture's own
        # user (also created via UserFactory) on User.username's unique constraint.
        # Give this second user an explicit unique one.
        outsider = UserFactory(username="outsider-user")
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001")])

        with pytest.raises(CommandError, match="not an active member"):
            call_command(
                "import_historical_project_data",
                str(xlsx_path),
                "--workspace",
                workspace.slug,
                "--created-by",
                outsider.email,
            )

    def test_unknown_workspace_raises_command_error(self, workspace_and_member, tmp_path):
        _workspace, user = workspace_and_member
        xlsx_path = _build_workbook(tmp_path / "source.xlsx", [_row("HD2024-001&PRJ2024-001")])

        with pytest.raises(CommandError, match="Workspace not found"):
            call_command(
                "import_historical_project_data",
                str(xlsx_path),
                "--workspace",
                "does-not-exist",
                "--created-by",
                user.email,
            )

    def test_missing_file_raises_command_error(self, workspace_and_member, tmp_path):
        workspace, user = workspace_and_member

        with pytest.raises(CommandError, match="File not found"):
            call_command(
                "import_historical_project_data",
                str(tmp_path / "does-not-exist.xlsx"),
                "--workspace",
                workspace.slug,
                "--created-by",
                user.email,
            )
