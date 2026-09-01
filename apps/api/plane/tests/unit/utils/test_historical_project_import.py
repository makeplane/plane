# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

"""
Unit tests for plane/utils/historical_project_import.py's pure parsing/coercion
helpers (Phase 3 of the custom-fields roadmap). These need no database: the
module has no Django import, so this exercises it directly against real
openpyxl worksheets.
"""

import datetime as dt
from decimal import Decimal

import openpyxl
import pytest

from plane.utils.historical_project_import import (
    coerce_date,
    coerce_dropdown,
    coerce_number,
    is_row_blank,
    parse_row,
    sanitize_project_text,
    validate_headers,
)

FIELD_SPECS = [
    {"name": "合同号&项目号", "field_type": "text"},
    {"name": "签约登记日期", "field_type": "date"},
    {"name": "税率（%）", "field_type": "number", "is_percent": True},
    {"name": "项目类别", "field_type": "dropdown", "options": ["A", "B", "C", "D"]},
]


def _worksheet(rows):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for col_index, spec in enumerate(FIELD_SPECS, start=1):
        worksheet.cell(row=4, column=col_index, value=spec["name"])
    for row_offset, row_values in enumerate(rows):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=1 + row_offset, column=col_index, value=value)
    return worksheet


@pytest.mark.unit
class TestSanitizeProjectText:
    def test_strips_forbidden_chars(self):
        assert sanitize_project_text("某零售(集团)公司&子公司") == "某零售集团公司子公司"

    def test_none_and_blank_return_none(self):
        assert sanitize_project_text(None) is None
        assert sanitize_project_text("   ") is None

    def test_truncates_to_max_length(self):
        assert sanitize_project_text("abcdef", max_length=3) == "abc"


@pytest.mark.unit
class TestIsRowBlank:
    def test_all_none_is_blank(self):
        assert is_row_blank([None, None, None])

    def test_whitespace_only_is_blank(self):
        assert is_row_blank(["  ", None, ""])

    def test_any_content_is_not_blank(self):
        assert not is_row_blank([None, "x", None])


@pytest.mark.unit
class TestCoerceNumber:
    def test_percent_number_format_multiplies_fraction(self):
        value, warning = coerce_number(0.13, "0%", is_percent=True)
        assert value == Decimal("13")
        assert warning is None

    def test_literal_percent_string_is_not_double_multiplied(self):
        value, warning = coerce_number("13%", None, is_percent=True)
        assert value == Decimal("13")
        assert warning is None

    def test_literal_percent_string_applies_even_when_not_is_percent(self):
        # Unlike number_format, a literal "%" in the cell's own text is an explicit
        # signal from the data itself, so it's honored regardless of is_percent.
        value, warning = coerce_number("13%", None, is_percent=False)
        assert value == Decimal("13")
        assert warning is None

    def test_percent_format_on_non_percent_field_is_not_multiplied_but_warns(self):
        # A field NOT marked is_percent (e.g. a money amount) whose cell happens to
        # carry percent formatting anyway must not be silently multiplied by 100:
        # that would corrupt a financial figure on a guess. It should come back
        # as the raw value with a warning instead.
        value, warning = coerce_number(0.13, "0%", is_percent=False)
        assert value == Decimal("0.13")
        assert warning is not None
        assert "is_percent" in warning

    def test_plain_number_passthrough(self):
        value, warning = coerce_number(128.5, "General")
        assert value == Decimal("128.5")

    def test_comma_thousands_separator(self):
        value, warning = coerce_number("1,234.56", "General")
        assert value == Decimal("1234.56")

    def test_unparseable_number_returns_warning(self):
        value, warning = coerce_number("not-a-number", "General")
        assert value is None
        assert warning is not None

    def test_blank_returns_none_no_warning(self):
        value, warning = coerce_number(None, "General")
        assert value is None
        assert warning is None


@pytest.mark.unit
class TestCoerceDate:
    def test_native_datetime_cell(self):
        value, warning = coerce_date(dt.datetime(2024, 5, 1))
        assert value == dt.date(2024, 5, 1)
        assert warning is None

    def test_string_date_common_formats(self):
        for text in ("2024-03-15", "2024/03/15", "2024.03.15"):
            value, warning = coerce_date(text)
            assert value == dt.date(2024, 3, 15)
            assert warning is None

    def test_unparseable_date_returns_warning(self):
        value, warning = coerce_date("not-a-date")
        assert value is None
        assert warning is not None


@pytest.mark.unit
class TestCoerceDropdown:
    def test_matching_option(self):
        value, warning = coerce_dropdown("A", ["A", "B", "C"])
        assert value == "A"
        assert warning is None

    def test_unmatched_option_returns_warning_not_value(self):
        value, warning = coerce_dropdown("Z", ["A", "B", "C"])
        assert value is None
        assert warning is not None

    def test_blank_returns_none_no_warning(self):
        value, warning = coerce_dropdown(None, ["A", "B", "C"])
        assert value is None
        assert warning is None


@pytest.mark.unit
class TestParseRow:
    def test_full_row_no_warnings(self):
        worksheet = _worksheet([["HD001&PRJ001", "2024-03-15", 0.13, "A"]])
        cell = worksheet.cell(row=3, column=3)
        cell.number_format = "0%"
        raw, coerced, warnings = parse_row(worksheet, 1, FIELD_SPECS)
        assert warnings == []
        assert coerced["合同号&项目号"] == "HD001&PRJ001"
        assert coerced["签约登记日期"] == dt.date(2024, 3, 15)
        assert coerced["项目类别"] == "A"

    def test_bad_cell_produces_warning_but_other_fields_still_parsed(self):
        worksheet = _worksheet([["HD001&PRJ001", "garbage-date", 0.13, "A"]])
        raw, coerced, warnings = parse_row(worksheet, 1, FIELD_SPECS)
        assert coerced["签约登记日期"] is None
        assert any("签约登记日期" in w for w in warnings)
        assert coerced["合同号&项目号"] == "HD001&PRJ001"  # unrelated fields unaffected


@pytest.mark.unit
class TestValidateHeaders:
    def test_matching_headers_return_no_mismatches(self):
        worksheet = _worksheet([])
        assert validate_headers(worksheet, header_row=4, field_specs=FIELD_SPECS) == []

    def test_mismatched_header_is_reported_with_column_letter(self):
        worksheet = _worksheet([])
        worksheet.cell(row=4, column=1, value="WRONG")
        mismatches = validate_headers(worksheet, header_row=4, field_specs=FIELD_SPECS)
        assert len(mismatches) == 1
        assert "column A" in mismatches[0]

    def test_source_header_overrides_name_for_matching(self):
        # Real-world case: DEFAULT_PROJECT_CUSTOM_FIELDS gives "税率（%）" a
        # source_header of "税率" because the source spreadsheet's literal column
        # header omits the unit-hint suffix added to the field's display name.
        specs = [{"name": "税率（%）", "field_type": "number", "source_header": "税率"}]
        worksheet = _worksheet([])
        worksheet.cell(row=4, column=1, value="税率")
        assert validate_headers(worksheet, header_row=4, field_specs=specs) == []

        worksheet.cell(row=4, column=1, value="税率（%）")  # the display name itself must NOT match
        mismatches = validate_headers(worksheet, header_row=4, field_specs=specs)
        assert len(mismatches) == 1
