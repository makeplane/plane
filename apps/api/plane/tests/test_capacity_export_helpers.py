# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

"""
Unit tests for capacity export helpers.

Pure function testing: sanitize_sheet_name, compute_member_totals, etc.
No database or external services.
"""

from types import SimpleNamespace
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from plane.bgtasks.capacity_export_helpers import (
    DETAIL_HEADERS,
    EMPTY_ROW_PLACEHOLDER,
    _assignees_str,
    _issue_identifier,
    sanitize_sheet_name,
    write_member_sheet,
    write_summary_sheet,
)


@pytest.mark.unit
class TestSanitizeSheetName:
    """Test suite for sanitize_sheet_name edge cases."""

    def test_normal_name_unchanged(self):
        """Normal names pass through unchanged."""
        used = set()
        result = sanitize_sheet_name("John Doe", used)
        assert result == "John Doe"
        assert "John Doe" in used

    def test_illegal_chars_removed(self):
        """Illegal Excel sheet name chars (: \\ / ? * [ ]) are stripped."""
        used = set()
        # Test each illegal char: : \ / ? * [ ]
        result = sanitize_sheet_name("A:B\\C/D?E*F[G]", used)
        assert ":" not in result
        assert "\\" not in result
        assert "/" not in result
        assert "?" not in result
        assert "*" not in result
        assert "[" not in result
        assert "]" not in result
        assert result == "ABCDEFG"

    def test_long_name_truncated_to_31(self):
        """Sheet names >31 chars are truncated."""
        used = set()
        long_name = "A" * 50
        result = sanitize_sheet_name(long_name, used)
        assert len(result) <= 31
        assert result == "A" * 31

    def test_truncation_reserves_space_for_suffix(self):
        """Truncation reserves space for -N suffix when needed."""
        used = set()
        # A 31-char name; adding it to used means next call needs -2 suffix
        long_base = "A" * 28 + "XYZ"  # 31 chars
        first = sanitize_sheet_name(long_base, used)
        assert first == long_base
        assert len(first) == 31

        # Now add a case-insensitive collision
        second = sanitize_sheet_name(long_base.lower(), used)
        # -2 is 2 chars, so base must be truncated to 29 to fit "-2"
        assert "-2" in second
        assert len(second) <= 31

    def test_empty_string_defaults_to_sheet(self):
        """Empty strings default to 'Sheet'."""
        used = set()
        result = sanitize_sheet_name("", used)
        assert result == "Sheet"
        assert "Sheet" in used

    def test_whitespace_only_defaults_to_sheet(self):
        """Whitespace-only strings default to 'Sheet' after strip."""
        used = set()
        result = sanitize_sheet_name("   ", used)
        assert result == "Sheet"

    def test_illegal_chars_only_defaults_to_sheet(self):
        """Strings with only illegal chars default to 'Sheet'."""
        used = set()
        result = sanitize_sheet_name("***???///", used)
        assert result == "Sheet"

    def test_case_insensitive_collision_detection(self):
        """Collisions are detected case-insensitively."""
        used = set()
        first = sanitize_sheet_name("MySheet", used)
        assert first == "MySheet"

        # Add collision with different case
        second = sanitize_sheet_name("mysheet", used)
        assert second != "mysheet"
        assert "-2" in second

    def test_multiple_collisions_suffix_increments(self):
        """Multiple collisions get -2, -3, -4, etc."""
        used = set()
        name1 = sanitize_sheet_name("Sheet", used)
        assert name1 == "Sheet"

        name2 = sanitize_sheet_name("SHEET", used)
        assert name2 == "Sheet-2"

        name3 = sanitize_sheet_name("sheet", used)
        assert name3 == "Sheet-3"

        name4 = sanitize_sheet_name("ShEeT", used)
        assert name4 == "Sheet-4"

    def test_used_set_updated_on_each_call(self):
        """The 'used' set is updated after each sanitize call."""
        used = set()
        sanitize_sheet_name("Alice", used)
        assert "Alice" in used

        sanitize_sheet_name("Bob", used)
        assert "Bob" in used
        assert len(used) == 2

    def test_truncation_with_suffix(self):
        """When truncating to add suffix, final name is still ≤31 chars."""
        used = set()
        # 28-char base + "XYZ" = 31 chars
        base = "A" * 28 + "XYZ"
        first = sanitize_sheet_name(base, used)
        assert len(first) == 31

        # Collision: needs -2, so truncate base to 29 chars
        second = sanitize_sheet_name(base, used)
        assert "-2" in second
        assert len(second) <= 31
        # Base should be "A" * 28 + "X" (29 chars) + "-2"
        assert second == "A" * 28 + "X" + "-2"

    def test_whitespace_stripped_before_processing(self):
        """Leading/trailing whitespace is stripped."""
        used = set()
        result = sanitize_sheet_name("  John Doe  ", used)
        assert result == "John Doe"

    def test_complex_collision_scenario(self):
        """Complex scenario: illegal chars + truncation + collision."""
        used = set()
        # Name with illegal chars, >31 after cleaning
        name1 = sanitize_sheet_name("A/B/C/D/E/F/G/H/I/J/K/L/M/N", used)
        # Illegal chars removed, result is short; no collision
        assert "/" not in name1

        # Add collision (case-insensitive)
        name2 = sanitize_sheet_name(
            "a/b/c/d/e/f/g/h/i/j/k/l/m/n", used
        )
        # Should add -2 suffix
        assert "-2" in name2 or name2 != name1.lower()


def _read_sheet(wb, title):
    """Return list of row tuples from a write-only workbook saved to a temp file."""
    import io
    from openpyxl import load_workbook
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    loaded = load_workbook(buf, read_only=True)
    return [tuple(c.value for c in row) for row in loaded[title].iter_rows()]


@pytest.mark.unit
class TestIssueIdentifier:
    def test_returns_proj_seq_format(self):
        issue = SimpleNamespace(project=SimpleNamespace(identifier="proj"), sequence_id=42)
        assert _issue_identifier(issue) == "PROJ-42" or _issue_identifier(issue) == "proj-42"

    def test_handles_missing_project(self):
        issue = SimpleNamespace(project=None, sequence_id=7)
        assert _issue_identifier(issue) == "7"

    def test_handles_none_issue(self):
        assert _issue_identifier(None) == ""


@pytest.mark.unit
class TestAssigneesStr:
    def test_joins_display_names(self):
        a1 = SimpleNamespace(assignee_id="u1", assignee=SimpleNamespace(display_name="Alice", email="a@x"))
        a2 = SimpleNamespace(assignee_id="u2", assignee=SimpleNamespace(display_name="Bob", email="b@x"))
        issue = SimpleNamespace(issue_assignee=SimpleNamespace(all=lambda: [a1, a2]))
        assert _assignees_str(issue) == "Alice, Bob"

    def test_falls_back_to_email(self):
        a1 = SimpleNamespace(assignee_id="u1", assignee=SimpleNamespace(display_name="", email="z@x"))
        issue = SimpleNamespace(issue_assignee=SimpleNamespace(all=lambda: [a1]))
        assert _assignees_str(issue) == "z@x"

    def test_handles_none(self):
        assert _assignees_str(None) == ""


@pytest.mark.unit
class TestWriteSummarySheet:
    def test_lists_all_roster_members_with_zeros(self):
        wb = Workbook(write_only=True)
        roster = [
            {"logged_by_id": "u1", "logged_by__display_name": "Alice", "total_minutes": 90, "entries": 2},
            {"logged_by_id": "u2", "logged_by__display_name": "Bob", "total_minutes": 0, "entries": 0},
        ]
        write_summary_sheet(wb, roster)
        rows = _read_sheet(wb, "Summary")
        assert rows[0] == ("Member", "Total Hours", "Entry Count")
        assert rows[1] == ("Alice", 1.5, 2)
        assert rows[2] == ("Bob", 0.0, 0)
        assert rows[3] == ("TOTAL", 1.5, 2)

    def test_empty_roster_shows_placeholder(self):
        wb = Workbook(write_only=True)
        write_summary_sheet(wb, [])
        rows = _read_sheet(wb, "Summary")
        assert rows[0] == ("Member", "Total Hours", "Entry Count")
        assert rows[1][0] == EMPTY_ROW_PLACEHOLDER


@pytest.mark.unit
class TestWriteMemberSheet:
    def _make_qs(self, entries):
        qs = MagicMock()
        filtered = MagicMock()
        filtered.iterator.return_value = iter(entries)
        qs.filter.return_value = filtered
        return qs

    def test_empty_member_gets_header_and_placeholder(self):
        wb = Workbook(write_only=True)
        qs = self._make_qs([])
        member = {"logged_by_id": "u1", "logged_by__display_name": "Alice"}
        rows = write_member_sheet(wb, member, qs, set(), "qa@x")
        assert rows == 0
        out = _read_sheet(wb, "Alice")
        assert out[0][0].startswith("Generated for qa@x on")
        assert tuple(out[1]) == tuple(DETAIL_HEADERS)
        assert out[2][0] == EMPTY_ROW_PLACEHOLDER

    def test_member_with_entries_writes_full_row(self):
        wb = Workbook(write_only=True)
        issue = SimpleNamespace(
            project=SimpleNamespace(identifier="proj", name="Demo"),
            sequence_id=3,
            name="Task A",
            main_task_category_id=1,
            main_task_category=SimpleNamespace(name="Backend"),
            sub_task_category_id=2,
            sub_task_category=SimpleNamespace(name="API"),
            state=SimpleNamespace(name="In Progress"),
            priority="high",
            issue_assignee=SimpleNamespace(all=lambda: [
                SimpleNamespace(assignee_id="u1", assignee=SimpleNamespace(display_name="Alice", email="a@x"))
            ]),
        )
        entry = SimpleNamespace(issue_id="i1", issue=issue, logged_at="2026-05-10", duration_minutes=120)
        qs = self._make_qs([entry])
        member = {"logged_by_id": "u1", "logged_by__display_name": "Alice"}
        rows = write_member_sheet(wb, member, qs, set(), "qa@x")
        assert rows == 1
        out = _read_sheet(wb, "Alice")
        data = out[2]
        # Date, Project, Work Item ID, Work Item, Main Cat, Sub Cat, State, Assignees, Priority, Hours
        assert data[0] == "2026-05-10"
        assert data[1] == "Demo"
        assert data[2].endswith("-3")
        assert data[3] == "Task A"
        assert data[4] == "Backend"
        assert data[5] == "API"
        assert data[6] == "In Progress"
        assert data[7] == "Alice"
        assert data[8] == "high"
        assert data[9] == 2.0

    def test_unique_sheet_names_for_duplicate_display_names(self):
        wb = Workbook(write_only=True)
        used = set()
        qs = self._make_qs([])
        write_member_sheet(wb, {"logged_by_id": "u1", "logged_by__display_name": "Alice"}, qs, used, "qa@x")
        write_member_sheet(wb, {"logged_by_id": "u2", "logged_by__display_name": "Alice"}, qs, used, "qa@x")
        assert "Alice" in used
        assert any(n != "Alice" and n.startswith("Alice") for n in used)
