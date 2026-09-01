# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.
#
# Internal addition (not part of upstream makeplane/plane): Phase 3 of the custom
# fields roadmap. Phase 1/2 built the field engine and the standard 18 project-side
# fields (those columns of the source contract/delivery tracking spreadsheet that
# describe the project itself, see apps/api/plane/db/default_data/
# project_custom_fields.py); this backfills real historical rows into Projects
# carrying those fields, while the contract-side columns (F 合同号, G 签约登记日期,
# H 合同净额, I 税率, J 合同占比) are routed into the new Contract /
# ContractProject models instead (see apps/api/plane/db/models/contract.py and
# docs/internal-contract-project-relationship.md). Column A (the legacy
# "合同号&项目号" composite identifier) is deliberately ignored -- per the
# 2026-09-01 business rule, that composite is no longer a data field and is
# rebuilt from Contract.contract_no + Project's "项目序号" wherever it needs to
# surface. Columns beyond W (the per-month financial matrix etc.) are explicitly
# out of scope for this phase -- see that roadmap's "Not building" section.
#
# Usage:
#   python manage.py import_historical_project_data <xlsx_path> \
#       --workspace <workspace-slug> --created-by <user-email> [--dry-run]
#       [--start-row 7] [--header-row 4] [--sheet Sheet1]
#
# Idempotent: a row whose "项目序号" value already exists (as an is_unique_key
# custom field value in this workspace, matching how the live API enforces
# uniqueness -- see ProjectCustomFieldValueSerializer.validate()) is skipped
# rather than re-imported, so re-running after fixing upstream data or extending
# the sheet is safe. --dry-run runs every validation and DB read the real
# import would, then rolls back all writes at the end.

# Python imports
import os
import re

# Third-party imports
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

# Module imports
from plane.app.permissions import ROLE
from plane.app.serializers import ProjectSerializer
from plane.db.default_data.project_custom_fields import DEFAULT_PROJECT_CUSTOM_FIELDS, seed_default_custom_fields
from plane.db.models import (
    Contract,
    ContractProject,
    DEFAULT_STATES,
    Project,
    ProjectCustomField,
    ProjectCustomFieldValue,
    ProjectMember,
    ProjectUserProperty,
    State,
    User,
    Workspace,
    WorkspaceMember,
)
from plane.utils.historical_project_import import (
    coerce_cell,
    is_row_blank,
    parse_row,
    sanitize_project_text,
    validate_headers,
)

# The single field DEFAULT_PROJECT_CUSTOM_FIELDS marks is_unique_key=True -- looked
# up by the flag rather than by hardcoded name, so changing the project-side
# unique-key field name in the future doesn't require touching this command.
UNIQUE_KEY_FIELD_NAME = next(
    spec["name"] for spec in DEFAULT_PROJECT_CUSTOM_FIELDS if spec.get("is_unique_key")
)
PROJECT_NAME_CANDIDATE_FIELDS = ("客户项目名称", "公司项目名称")

# Source spreadsheet columns that route into the new Contract / ContractProject
# models (rather than into ProjectCustomFieldValue). Listed as (column_letter,
# spec_dict) so the column letter can be reported in error messages. The
# spec_dict carries enough info to call coerce_cell() without re-reading the
# DEFAULT_PROJECT_CUSTOM_FIELDS list (which doesn't contain these fields
# anymore -- they're not project-side data).
_CONTRACT_SIDE_COLUMNS = (
    {"column": "F", "name": "合同号", "field_type": "text", "target": "Contract.contract_no"},
    {"column": "G", "name": "签约登记日期", "field_type": "date", "target": "Contract.sign_date"},
    {
        "column": "H",
        "name": "合同净额/不含第三方（人民币万元）",
        "field_type": "number",
        "target": "Contract.total_amount",
    },
    {
        "column": "I",
        "name": "税率（%）",
        "field_type": "number",
        "source_header": "税率",
        "is_percent": True,
        "target": "Contract.tax_rate",
    },
    {
        "column": "J",
        "name": "合同占比（%）",
        "field_type": "number",
        "source_header": "合同占比",
        "is_percent": True,
        "target": "ContractProject.allocation_ratio",
    },
)

# Pre-Phase-A source spreadsheet column A was a data column ("合同号&项目号"); the
# 2026-09-01 business rule retired it entirely. We only need its column letter
# for the header-presence check below; its contents are NEVER coerced, NEVER
# written, and any value is accepted without warning.


def _coerce_contract_cell(spec, raw_value, number_format):
    """
    Same coercion as parse_row does for a field_spec entry, but reused here for the
    contract-side columns (which live outside DEFAULT_PROJECT_CUSTOM_FIELDS now
    that they moved to Contract / ContractProject). Kept inline rather than
    imported out of historical_project_import to keep that module's surface
    unchanged for its existing tests.

    Excel-specific round-trip: the source spreadsheet's "合同号" column (F) holds
    strings like "5763-5", but Excel auto-parses anything that looks like YYYY-M
    into a datetime object at open time. Without the early returns below,
    coerce_text() would format such a value as ISO ("5763-05-01 00:00:00"),
    breaking both uniqueness and human display. The (year, month) reconstruction
    only fires when the time-of-day is exactly midnight, which is Excel's
    default for parsed dates with no time component -- a real datetime value in
    a text cell would have non-zero time fields and pass through unchanged.
    Int inputs are a smaller sibling: Excel silently drops a trailing "-NN"
    suffix when the cell ends up typed as a number, but at least the integer
    part is preserved (we return it as a string rather than re-adding the lost
    suffix, which we can't recover from a single column).
    """
    if spec["field_type"] == "text":
        import datetime as _dt
        if isinstance(raw_value, _dt.datetime) and raw_value.hour == 0 and raw_value.minute == 0 and raw_value.second == 0 and raw_value.microsecond == 0:
            return f"{raw_value.year}-{raw_value.month}", None
        if isinstance(raw_value, _dt.date) and not isinstance(raw_value, _dt.datetime):
            return f"{raw_value.year}-{raw_value.month}", None
        if isinstance(raw_value, int):
            return str(raw_value), None
    return coerce_cell(
        spec["field_type"],
        raw_value,
        number_format=number_format,
        options=spec.get("options"),
        is_percent=spec.get("is_percent", False),
    )


class Command(BaseCommand):
    help = (
        "Import historical project rows from the source contract/delivery tracking "
        "spreadsheet: one Project per row, populated into the 18 project-side custom "
        "fields seed_default_custom_fields builds, plus one Contract row per distinct "
        "合同号 and one ContractProject link row per Project (carrying 合同占比). The "
        "retired composite \"合同号&项目号\" column (A) is ignored as data."
    )

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to the source .xlsx file")
        parser.add_argument("--workspace", required=True, dest="workspace_slug", help="Target workspace slug")
        parser.add_argument(
            "--created-by",
            required=True,
            dest="created_by_email",
            help="Email of the (already-a-workspace-member) user to attribute imported projects to",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Run every validation and DB read, then roll back all writes instead of committing",
        )
        parser.add_argument("--start-row", type=int, default=7, help="First data row (default 7)")
        parser.add_argument("--header-row", type=int, default=4, help="Row holding the A-W field names (default 4)")
        parser.add_argument("--sheet", type=str, default=None, help="Sheet name (default: the workbook's first sheet)")

    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        if not os.path.isfile(xlsx_path):
            raise CommandError(f"File not found: {xlsx_path}")

        try:
            workspace = Workspace.objects.get(slug=options["workspace_slug"])
        except Workspace.DoesNotExist as exc:
            raise CommandError(f"Workspace not found: {options['workspace_slug']}") from exc

        try:
            # User.save() lowercases email before storing (see plane/db/models/user.py),
            # so match case-insensitively here too; a case-sensitive lookup would fail
            # to find a real account whenever the operator types it with any capitals.
            created_by = User.objects.get(email=options["created_by_email"].strip().lower())
        except User.DoesNotExist as exc:
            raise CommandError(f"User not found: {options['created_by_email']}") from exc

        if created_by.is_bot or not created_by.is_active:
            raise CommandError(
                f"{options['created_by_email']} is a bot account or globally deactivated; "
                "refusing to attribute imported projects to it."
            )

        if not WorkspaceMember.objects.filter(workspace=workspace, member=created_by, is_active=True).exists():
            raise CommandError(
                f"{options['created_by_email']} is not an active member of workspace "
                f"{workspace.slug}; refusing to attribute imported projects to a non-member."
            )

        workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        worksheet = workbook[options["sheet"]] if options["sheet"] else workbook.worksheets[0]

        # Phase A: pass header_row so validate_headers / parse_row do
        # header-name lookup instead of the pre-Phase-A positional behaviour. The
        # 2026-09-01 retirement of column A ("合同号&项目号") means
        # DEFAULT_PROJECT_CUSTOM_FIELDS no longer mirrors the source spreadsheet's
        # A-W column order, so the positional path would fail to find each spec.
        header_row = options["header_row"]
        mismatches = validate_headers(worksheet, header_row, DEFAULT_PROJECT_CUSTOM_FIELDS)
        if mismatches:
            raise CommandError(
                "Header row does not match the 18 seeded project-side custom fields. "
                "Fix the source file or pass --header-row, then re-run:\n" + "\n".join(mismatches)
            )

        # Build a one-time header->column map for the contract-side columns (which
        # live outside DEFAULT_PROJECT_CUSTOM_FIELDS now that they moved to
        # Contract / ContractProject). Source_header preferred where it differs
        # from the field's display name (e.g. "税率" in the source vs "税率（%）"
        # in the Plane UI).
        source_header_map = {}
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=header_row, column=col).value
            if value is not None:
                text = str(value).strip()
                if text and text not in source_header_map:
                    source_header_map[text] = col
        contract_columns = []
        for spec in _CONTRACT_SIDE_COLUMNS:
            key = spec.get("source_header", spec["name"])
            col = source_header_map.get(key)
            if col is None and spec.get("name") != key:
                col = source_header_map.get(spec["name"])
            contract_columns.append({**spec, "xlsx_col": col})
        # The retired column A presence-check: it's data we ignore, but if it's
        # missing entirely the user probably pointed at the wrong header row.
        retired_col = source_header_map.get("合同号&项目号")
        if retired_col is None:
            self.stderr.write(
                self.style.WARNING(
                    f"Header row {header_row} has no '合同号&项目号' column -- it will be "
                    "ignored if present (per the 2026-09-01 business rule). Continuing."
                )
            )

        dry_run = options["dry_run"]
        stats = {"created": 0, "skipped_blank": 0, "skipped_no_key": 0, "skipped_duplicate": 0, "skipped_error": 0}
        field_warning_count = 0
        seen_keys_this_run = set()
        seen_contracts_this_run = set()  # Contract.contract_no -> Contract (for get_or_create idempotency within one import run)

        for row_idx in range(options["start_row"], worksheet.max_row + 1):
            raw_values, coerced, row_warnings = parse_row(
                worksheet, row_idx, DEFAULT_PROJECT_CUSTOM_FIELDS, header_row=header_row
            )

            if is_row_blank(raw_values):
                stats["skipped_blank"] += 1
                continue

            # Read the contract-side cells directly (not via parse_row, since those
            # columns are not in DEFAULT_PROJECT_CUSTOM_FIELDS anymore).
            contract_coerced = {}
            contract_warnings = []
            for spec in contract_columns:
                if spec["xlsx_col"] is None:
                    contract_coerced[spec["name"]] = None
                    continue
                cell = worksheet.cell(row=row_idx, column=spec["xlsx_col"])
                value, warning = _coerce_contract_cell(spec, cell.value, cell.number_format)
                contract_coerced[spec["name"]] = value
                if warning:
                    contract_warnings.append(f"{spec['name']}: {warning}")

            unique_key = coerced.get(UNIQUE_KEY_FIELD_NAME)
            if not unique_key:
                stats["skipped_no_key"] += 1
                self.stderr.write(self.style.WARNING(f"Row {row_idx}: skipped, missing {UNIQUE_KEY_FIELD_NAME}"))
                continue

            contract_no = contract_coerced.get("合同号")
            # A row with no 合同号 (e.g. "暂无" / "待签约" placeholder) is still
            # importable as a Project -- it just won't link to a Contract. We
            # surface this with a row-level warning rather than aborting.
            if not contract_no:
                self.stderr.write(
                    self.style.WARNING(
                        f"Row {row_idx} ({unique_key!r}): no 合同号 -- project will be imported without a Contract link"
                    )
                )
                field_warning_count += 1

            try:
                with transaction.atomic():
                    # Mirrors ProjectCustomFieldValueViewSet.partial_update()'s lock:
                    # the duplicate check below is a separate SELECT from this row's
                    # eventual INSERT, and no DB constraint can enforce workspace-wide
                    # uniqueness directly (the value lives on a per-project row). A
                    # Postgres advisory lock keyed on (workspace, value) serializes
                    # only writers racing for the *same* value -- other rows in this
                    # same run, another concurrent run of this command, and any live
                    # API write through that endpoint -- against each other.
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "SELECT pg_advisory_xact_lock(hashtext(%s))", [f"{workspace.id}:{unique_key}"]
                        )

                    is_duplicate = unique_key in seen_keys_this_run or ProjectCustomFieldValue.objects.filter(
                        workspace=workspace,
                        custom_field__is_unique_key=True,
                        custom_field__deleted_at__isnull=True,
                        value_text=unique_key,
                        deleted_at__isnull=True,
                    ).exists()
                    if is_duplicate:
                        stats["skipped_duplicate"] += 1
                        self.stderr.write(self.style.WARNING(f"Row {row_idx}: skipped, duplicate {unique_key!r}"))
                        continue

                    project = self._create_project_for_row(workspace, created_by, coerced, row_idx)
                    self._write_field_values(workspace, project, created_by, coerced)

                    contract = None
                    if contract_no:
                        contract = self._get_or_create_contract(
                            workspace, created_by, contract_coerced, seen_contracts_this_run
                        )
                        self._link_contract_to_project(
                            workspace, project, created_by, contract, contract_coerced
                        )

                    if dry_run:
                        transaction.set_rollback(True)
            except Exception as exc:  # noqa: BLE001 -- one bad row must not abort the whole import
                stats["skipped_error"] += 1
                self.stderr.write(self.style.ERROR(f"Row {row_idx} ({unique_key!r}): failed, {exc}"))
                continue

            if unique_key in seen_keys_this_run:
                # The "continue" for the duplicate branch above lands here too (it
                # exits the `with` block cleanly, not via exception), so re-check
                # before counting this as a fresh create.
                continue

            seen_keys_this_run.add(unique_key)
            stats["created"] += 1
            field_warning_count += len(row_warnings) + len(contract_warnings)
            for warning in row_warnings + contract_warnings:
                self.stderr.write(self.style.WARNING(f"Row {row_idx} ({unique_key!r}): {warning}"))

        self.stdout.write(
            self.style.SUCCESS(
                f"{'[DRY RUN] ' if dry_run else ''}Created {stats['created']} project(s). "
                f"Skipped: {stats['skipped_blank']} blank, {stats['skipped_no_key']} missing unique key, "
                f"{stats['skipped_duplicate']} duplicate, {stats['skipped_error']} error. "
                f"{field_warning_count} field-level warning(s) (see above)."
            )
        )

    def _create_project_for_row(self, workspace, created_by, coerced, row_idx):
        raw_name = next(
            (coerced.get(field_name) for field_name in PROJECT_NAME_CANDIDATE_FIELDS if coerced.get(field_name)),
            None,
        )
        # "-" is forbidden in Project.name (Project.FORBIDDEN_IDENTIFIER_CHARS_PATTERN),
        # so the fallback uses "_" as a separator, matching the existing convention in
        # apps/api/plane/bgtasks/dummy_data_task.py's create_project().
        fallback_key = sanitize_project_text(coerced[UNIQUE_KEY_FIELD_NAME], max_length=40) or "unnamed"
        name = sanitize_project_text(raw_name) or f"导入项目_{fallback_key}"
        name = self._unique_project_name(workspace, name)
        identifier = self._generate_identifier(workspace, name, row_idx)

        serializer = ProjectSerializer(data={"name": name, "identifier": identifier}, context={"workspace_id": workspace.id})
        serializer.is_valid(raise_exception=True)
        project = serializer.save()
        # ProjectSerializer.create() -> Project.objects.create() -> Model.save() reads
        # created_by from crum's request-thread-local, which is empty in a management
        # command; .update() bypasses save() entirely so it always lands regardless.
        Project.objects.filter(pk=project.pk).update(created_by=created_by)

        member = ProjectMember(project=project, workspace=workspace, member=created_by, role=ROLE.ADMIN.value)
        member.save()
        ProjectMember.objects.filter(pk=member.pk).update(created_by=created_by)
        # ProjectMember.save() itself creates a ProjectUserProperty row as a side
        # effect (see that model's save() override); same crum-less created_by gap
        # applies there too, so fix it up the same way.
        ProjectUserProperty.objects.filter(project=project, user=created_by).update(created_by=created_by)

        State.objects.bulk_create(
            [
                State(
                    name=state["name"],
                    color=state["color"],
                    project=project,
                    sequence=state["sequence"],
                    workspace=workspace,
                    group=state["group"],
                    default=state.get("default", False),
                    created_by=created_by,
                )
                for state in DEFAULT_STATES
            ]
        )

        seed_default_custom_fields(project, created_by=created_by)
        return project

    def _write_field_values(self, workspace, project, created_by, coerced):
        fields_by_name = {
            field.name: field for field in ProjectCustomField.objects.filter(project=project).prefetch_related("options")
        }
        value_rows = []
        for spec in DEFAULT_PROJECT_CUSTOM_FIELDS:
            value = coerced.get(spec["name"])
            if value is None:
                continue
            field = fields_by_name[spec["name"]]
            kwargs = {"project": project, "custom_field": field, "workspace_id": workspace.id, "created_by": created_by}
            if spec["field_type"] == "number":
                kwargs["value_decimal"] = value
            elif spec["field_type"] == "text":
                kwargs["value_text"] = value
            elif spec["field_type"] == "date":
                kwargs["value_date"] = value
            elif spec["field_type"] == "dropdown":
                option = next((opt for opt in field.options.all() if opt.name == value), None)
                if option is None:
                    # parse_row already validated the value against spec["options"], so
                    # this only fires if seeding somehow diverged from that list.
                    continue
                kwargs["value_option"] = option
            else:
                continue
            value_rows.append(ProjectCustomFieldValue(**kwargs))

        if value_rows:
            ProjectCustomFieldValue.objects.bulk_create(value_rows)

    def _get_or_create_contract(self, workspace, created_by, contract_coerced, seen_contracts_this_run):
        """
        Looks up an existing Contract for (workspace, contract_no) or creates a new one.
        Two layers of idempotency within one import run:

        1. The seen_contracts_this_run dict caches the row -> Contract mapping so a
           repeated contract_no across rows only hits the DB once per row (still a
           get_or_create in case another import session wrote it).
        2. Contract has a DB UniqueConstraint on (workspace, contract_no), so
           concurrent writers across import runs and live-API paths can't create
           duplicates. Unlike ProjectCustomFieldValue (which lives on a per-project
           row and cannot express workspace-wide uniqueness as a DB constraint,
           hence the advisory lock around its import path above), Contract is a
           real shared table -- the DB constraint is the source of truth, no
           advisory lock is needed here. Django's get_or_create does swallow
           IntegrityError and retry the SELECT under race conditions, so even
           concurrent writers within the same import run converge to one row
           without raising.

        Fields written: contract_no (required, from "合同号" column), sign_date
        ("签约登记日期"), total_amount ("合同净额"), tax_rate ("税率"). 合同占比 is
        NOT a Contract attribute -- it lives on ContractProject.allocation_ratio
        and is written by _link_contract_to_project below.
        """
        contract_no = contract_coerced["合同号"]
        if contract_no in seen_contracts_this_run:
            return seen_contracts_this_run[contract_no]
        contract, _created = Contract.objects.get_or_create(
            workspace=workspace,
            contract_no=contract_no,
            defaults={
                "sign_date": contract_coerced.get("签约登记日期"),
                "total_amount": contract_coerced.get("合同净额/不含第三方（人民币万元）"),
                "tax_rate": contract_coerced.get("税率（%）"),
                "created_by": created_by,
            },
        )
        # For an existing Contract, do NOT mutate its sign_date / total_amount /
        # tax_rate from later rows in the same file -- first-writer wins, matching
        # the source spreadsheet's de-facto behaviour (each contract's number
        # appears in multiple rows, but its financial fields are only meaningful
        # the first time). If the values diverge across rows that's already a
        # data-quality warning to surface separately.
        seen_contracts_this_run[contract_no] = contract
        return contract

    def _link_contract_to_project(self, workspace, project, created_by, contract, contract_coerced):
        """
        Idempotent join-row creation. ContractProject already has a DB
        UniqueConstraint on (contract, project), so get_or_create handles
        both "first row for this (contract, project) pair" and "row already
        linked by a prior import run". No advisory lock needed here -- the row
        uniqueness is a hard DB constraint, unlike the ProjectCustomFieldValue
        uniqueness above.
        """
        allocation_ratio = contract_coerced.get("合同占比（%）")
        link, _created = ContractProject.objects.get_or_create(
            contract=contract,
            project=project,
            workspace=workspace,
            defaults={
                "allocation_ratio": allocation_ratio,
                "created_by": created_by,
            },
        )
        if not _created and allocation_ratio is not None and link.allocation_ratio != allocation_ratio:
            # The same Contract-Project pair seen again with a different
            # allocation_ratio -- take the first occurrence as the authoritative
            # one and surface the discrepancy. Don't silently overwrite: that
            # would either quietly pick a wrong value or write a value the
            # operator never reviewed.
            self.stderr.write(
                self.style.WARNING(
                    f"Project {project.name!r} already linked to {contract.contract_no!r} with "
                    f"allocation_ratio={link.allocation_ratio}; ignoring new value {allocation_ratio}"
                )
            )

    def _unique_project_name(self, workspace, base_name):
        # "-" is forbidden in Project.name, so "_" separates the disambiguating suffix.
        # base_name can be up to 255 chars (sanitize_project_text's default
        # max_length): truncate IT, not just the final string, to leave room for the
        # suffix -- f"{base_name}_{suffix}"[:255] alone would silently truncate the
        # suffix away whenever base_name is already >= 255 chars, producing the same
        # colliding string on every iteration and looping forever.
        name = base_name
        suffix = 2
        while Project.objects.filter(workspace=workspace, name=name).exists():
            suffix_text = f"_{suffix}"
            name = f"{base_name[: 255 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        return name

    def _generate_identifier(self, workspace, name, row_idx):
        letters = "".join(ch for ch in name.upper() if ch.isalnum() and ch.isascii())
        base = letters[:8] or "PRJ"
        # Start from row_idx instead of 1: this dataset is Chinese-language, so many
        # rows fall back to the same short base ("PRJ", or one leftover Latin letter
        # from a name like "内部项目名A"). Starting every row's scan at suffix=1 would
        # make row N re-check ~N already-taken candidates before finding a free one
        # (O(rows^2) queries total). row_idx is unique per row and increases through
        # the file, so f"{base}{row_idx}" is already very likely free on the first try.
        suffix = row_idx
        while True:
            candidate = f"{base}{suffix}"[:12]
            if not re.match(Project.FORBIDDEN_IDENTIFIER_CHARS_PATTERN, candidate) and not Project.objects.filter(
                workspace=workspace, identifier=candidate
            ).exists():
                return candidate
            suffix += 1
